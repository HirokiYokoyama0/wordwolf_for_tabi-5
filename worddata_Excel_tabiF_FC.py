import openpyxl
import random

def create_word_TF(GENRUNUM):

    workbook = openpyxl.load_workbook('ワードウルフ単語_TF2.xlsx')
    sheet = workbook["General"]
    max_row_num = sheet.max_row
    word_data = []
    qest_data = []

    #word取得
    for j in range(2,13):  #word
        word = sheet.cell(row=GENRUNUM+1, column=j).value
        if word is None:
            break
        else:
            word_data.append(word)

    
    for j in range(14,19): # 5質問例
        qest = sheet.cell(row=GENRUNUM+1, column=j).value
        if  qest is None:
            qest_data.append("")
        else:
            qest_data.append(qest)


    selected_word = random.sample(word_data,2)
    #print("selected_word ==>",selected_word)
    
    return selected_word,qest_data

  
def check_genre():

    workbook = openpyxl.load_workbook('ワードウルフ単語_TF2.xlsx')
    sheet = workbook["General"]
    max_row_num = sheet.max_row
    genre_data = []

    
    for k in range(2,22):##ここは可変にすべき
        genre = sheet.cell(row=k, column=1).value
        genre_data.append(genre)

    return genre_data







