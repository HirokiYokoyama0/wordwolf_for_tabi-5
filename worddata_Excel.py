import openpyxl
import random

def create_word():

  workbook = openpyxl.load_workbook('ワードウルフ_質問データベース_No1.xlsx')

  sheet = workbook["General"]
  max_row_num = sheet.max_row
  word_data = []

  for i in range(2,max_row_num+1):
    word1 = sheet.cell(row=i, column=1).value #word1
    word2 = sheet.cell(row=i, column=2).value #word2
    qest1 = sheet.cell(row=i, column=3).value #質問1
    qest2 = sheet.cell(row=i, column=4).value #質問2
    qest3 = sheet.cell(row=i, column=5).value #質問3

    #if cell_value not in suppliers:
    word_data.append([word1,word2,qest1,qest2,qest3])

  #print("max_row_num==>",max_row_num,len(word_data))
  word_num = random.randint(0,len(word_data)-1)
  #print("word_num-->",word_num)
  #print("word_data[0]-->",word_data[word_num])
  #print("word_data[7][0]-->",word_data[word_num][0])
  #print("word_data[7][1]-->",word_data[word_num][1])
  # 戻り地として、重複のない乱数列とデータ一覧を戻す。
  return word_data,max_row_num
  


# 重複のない乱数を生成する
def rand_ints_nodup(a, b, k):
  ns = []
  while len(ns) < k:
    n = random.randint(a, b)
    if not n in ns:
      ns.append(n)
  return ns
